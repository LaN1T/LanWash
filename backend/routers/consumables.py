import uuid
import io
from datetime import datetime, timedelta
from fastapi import APIRouter, HTTPException, Depends, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from core.limiter import limiter
from sqlalchemy import select, update, delete, func
from database import get_db
from models import (
    ConsumableRequest, ConsumableResponse, ServiceConsumableRequest,
    ServiceConsumableResponse, RefillRequest,
)
from db_models import Consumable, ServiceConsumable, Service, User, ConsumableRefillLog, ConsumableUsageLog
from services.auth_service import get_current_user, check_roles

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter(
    prefix="/api/consumables",
    tags=["consumables"],
)

@router.get("/", response_model=list[ConsumableResponse])
@limiter.limit("60/minute")
async def get_all_consumables(request: Request, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(['admin', 'washer']))):
    result = await db.execute(select(Consumable).order_by(Consumable.name.asc()))
    return result.scalars().all()

@router.get("/by-service/{service_id}", response_model=list[ServiceConsumableResponse])
@limiter.limit("60/minute")
async def get_consumables_by_service(request: Request, service_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(['admin', 'washer']))):
    result = await db.execute(select(ServiceConsumable).where(ServiceConsumable.serviceId == service_id).order_by(ServiceConsumable.consumableId.asc()))
    return result.scalars().all()

@router.post("/", response_model=ConsumableResponse)
@limiter.limit("10/minute")
async def create_consumable(request: Request, req: ConsumableRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(['admin']))):
    new_consumable = Consumable(id=str(uuid.uuid4()), name=req.name, unit=req.unit)
    db.add(new_consumable)
    await db.commit()
    await db.refresh(new_consumable)
    return new_consumable

@router.get("/alerts/low-stock", response_model=list[ConsumableResponse])
@limiter.limit("60/minute")
async def get_low_stock_alerts(request: Request, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(['admin', 'washer']))):
    result = await db.execute(
        select(Consumable).where(Consumable.currentStock < Consumable.minStock).order_by(Consumable.name.asc())
    )
    return result.scalars().all()


# ========== Excel экспорт / импорт (статичные пути — до динамических) ==========

def _create_workbook():
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl не установлен")
    return openpyxl.Workbook()


def _style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _sanitize_excel(val):
    """Prevent Excel formula injection by prefixing dangerous characters."""
    if isinstance(val, str) and val and val[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + val
    return val


@router.get("/export")
@limiter.limit("10/minute")
async def export_consumables(
    request: Request,
    date_from: str = None,
    date_to: str = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(check_roles(['admin', 'washer'])),
):
    """Экспорт отчёта по расходникам в Excel."""
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl не установлен")

    wb = _create_workbook()
    wb.remove(wb.active)

    # 1. Лист "Остатки"
    ws_stock = wb.create_sheet("Остатки")
    headers_stock = ["ID", "Название", "Ед. изм.", "Текущий запас", "Мин. запас", "Статус"]
    ws_stock.append(headers_stock)
    for cell in ws_stock[1]:
        _style_header(cell)

    result = await db.execute(select(Consumable).order_by(Consumable.name.asc()))
    for c in result.scalars().all():
        status = "Низкий" if c.currentStock < c.minStock else "В норме"
        ws_stock.append([_sanitize_excel(c.id), _sanitize_excel(c.name), _sanitize_excel(c.unit), c.currentStock, c.minStock, _sanitize_excel(status)])
    _auto_width(ws_stock)

    # 2. Лист "Пополнения"
    ws_refill = wb.create_sheet("Пополнения")
    headers_refill = ["Расходник", "Количество", "Было", "Стало", "Кем", "Дата"]
    ws_refill.append(headers_refill)
    for cell in ws_refill[1]:
        _style_header(cell)

    query_refill = select(ConsumableRefillLog).order_by(ConsumableRefillLog.timestamp.desc())
    if date_from:
        query_refill = query_refill.where(ConsumableRefillLog.timestamp >= date_from)
    if date_to:
        query_refill = query_refill.where(ConsumableRefillLog.timestamp <= date_to)

    refill_logs = (await db.execute(query_refill)).scalars().all()
    # Bulk load consumable names
    refill_cids = {log.consumableId for log in refill_logs}
    cons_names: dict[str, str] = {}
    if refill_cids:
        c_res = await db.execute(
            select(Consumable.id, Consumable.name).where(Consumable.id.in_(refill_cids))
        )
        cons_names = {cid: name for cid, name in c_res.all()}

    for log in refill_logs:
        name = cons_names.get(log.consumableId, log.consumableId)
        ws_refill.append([
            _sanitize_excel(name),
            log.amount,
            log.oldStock,
            log.newStock,
            _sanitize_excel(log.refilledBy),
            _sanitize_excel(log.timestamp),
        ])
    _auto_width(ws_refill)

    # 3. Лист "Расход"
    ws_usage = wb.create_sheet("Расход")
    headers_usage = ["Расходник", "Ед. изм.", "Использовано", "Период"]
    ws_usage.append(headers_usage)
    for cell in ws_usage[1]:
        _style_header(cell)

    query_usage = select(ConsumableUsageLog)
    if date_from:
        query_usage = query_usage.where(ConsumableUsageLog.timestamp >= date_from)
    if date_to:
        query_usage = query_usage.where(ConsumableUsageLog.timestamp <= date_to)
    usage_logs = (await db.execute(query_usage)).scalars().all()

    # Collect unique consumable IDs
    usage_cids = {log.consumableId for log in usage_logs}
    # Bulk load consumable names and units
    consumables_res = await db.execute(
        select(Consumable.id, Consumable.name, Consumable.unit)
        .where(Consumable.id.in_(usage_cids))
    )
    consumables_map = {cid: (name, unit) for cid, name, unit in consumables_res.all()}

    usage_sums: dict[str, float] = {}
    for log in usage_logs:
        usage_sums[log.consumableId] = usage_sums.get(log.consumableId, 0.0) + log.quantityUsed

    period_label = f"{date_from or '...'} - {date_to or '...'}"
    for cid, total in sorted(usage_sums.items(), key=lambda x: x[1], reverse=True):
        name, unit = consumables_map.get(cid, (cid, ""))
        ws_usage.append([_sanitize_excel(name), _sanitize_excel(unit), round(total, 2), _sanitize_excel(period_label)])
    _auto_width(ws_usage)

    # 4. Лист "Прогноз"
    ws_forecast = wb.create_sheet("Прогноз")
    headers_forecast = ["Расходник", "Ед. изм.", "Текущий запас", "Средний расход/день", "Хватит на (дн)", "Реком. закупка"]
    ws_forecast.append(headers_forecast)
    for cell in ws_forecast[1]:
        _style_header(cell)

    all_consumables = (await db.execute(select(Consumable))).scalars().all()
    thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()

    # Bulk load usage sums for all consumables in one query
    usage_sums_res = await db.execute(
        select(
            ConsumableUsageLog.consumableId,
            func.coalesce(func.sum(ConsumableUsageLog.quantityUsed), 0.0),
        )
        .where(ConsumableUsageLog.timestamp >= thirty_days_ago)
        .group_by(ConsumableUsageLog.consumableId)
    )
    usage_sums_map = {cid: float(total) for cid, total in usage_sums_res.all()}

    for c in all_consumables:
        total_used = usage_sums_map.get(c.id, 0.0)
        avg_daily = total_used / 30.0 if total_used > 0 else 0.0
        days_left = round(c.currentStock / avg_daily, 1) if avg_daily > 0 else "—"
        target = c.minStock * 3
        to_buy = max(0.0, target - c.currentStock)
        ws_forecast.append([
            _sanitize_excel(c.name), _sanitize_excel(c.unit), c.currentStock,
            round(avg_daily, 2), _sanitize_excel(str(days_left)), round(to_buy, 1)
        ])
    _auto_width(ws_forecast)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"consumables_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/import-template")
@limiter.limit("10/minute")
async def download_import_template(request: Request):
    """Скачать пустой шаблон Excel для импорта пополнений."""
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl не установлен")

    wb = _create_workbook()
    ws = wb.active
    ws.title = "Пополнения"
    headers = ["name", "amount"]
    ws.append(headers)
    for cell in ws[1]:
        _style_header(cell)
    ws.append(["Автошампунь", 10.0])
    ws.append(["Воск", 5.0])
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=consumables_import_template.xlsx"},
    )


MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024  # 5 MB

@router.post("/import-refills")
@limiter.limit("10/minute")
async def import_refills(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(check_roles(['admin'])),
):
    """Массовый импорт пополнений из Excel. Ожидаются колонки: name, amount"""
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl не установлен")

    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(413, "Файл слишком большой. Максимум 5 МБ")

    content = await file.read()
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(413, "Файл слишком большой. Максимум 5 МБ")
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Не удалось открыть файл: {e}")

    ws = wb.active
    if not ws:
        raise HTTPException(400, "Файл пуст")

    header_row = [cell.value for cell in ws[1]]
    header_lower = [str(h).lower().strip() if h else "" for h in header_row]

    name_idx = None
    amount_idx = None
    for i, h in enumerate(header_lower):
        if h in ("name", "название", "расходник", "наименование"):
            name_idx = i
        if h in ("amount", "количество", "кол-во", "колво", "сумма"):
            amount_idx = i

    if name_idx is None or amount_idx is None:
        raise HTTPException(400, "Колонки 'name' и 'amount' не найдены в первой строке")

    succeeded = 0
    failed = 0
    errors: list[str] = []
    processed = 0

    # Загружаем все расходники один раз для O(n) поиска
    all_consumables = (await db.execute(select(Consumable))).scalars().all()
    consumables_by_name = {c.name.lower(): c for c in all_consumables}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        amount_raw = row[amount_idx]
        if not name:
            continue
        processed += 1

        try:
            amount = float(amount_raw) if amount_raw is not None else 0.0
        except (ValueError, TypeError):
            failed += 1
            errors.append(f"Строка {processed + 1}: '{amount_raw}' не является числом")
            continue

        if amount <= 0:
            failed += 1
            errors.append(f"Строка {processed + 1}: количество должно быть > 0")
            continue

        consumable = consumables_by_name.get(name.lower())
        if not consumable:
            failed += 1
            errors.append(f"Строка {processed + 1}: расходник '{name}' не найден")
            continue

        old_stock = consumable.currentStock
        consumable.currentStock += amount
        db.add(ConsumableRefillLog(
            consumableId=consumable.id,
            amount=amount,
            oldStock=old_stock,
            newStock=consumable.currentStock,
            refilledBy=current_user.username,
            timestamp=datetime.now().isoformat(),
        ))
        succeeded += 1

    await db.commit()

    return {
        "processed": processed,
        "succeeded": succeeded,
        "failed": failed,
        "errors": errors[:20],
    }


# ========== Динамические пути ==========

@router.get("/{consumable_id}", response_model=ConsumableResponse)
@limiter.limit("60/minute")
async def get_consumable(request: Request, consumable_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin", "washer"]))):
    result = await db.execute(select(Consumable).where(Consumable.id == consumable_id))
    consumable = result.scalar_one_or_none()
    if not consumable:
        raise HTTPException(404, "Расходник не найден")
    return consumable

@router.put("/{consumable_id}", response_model=ConsumableResponse)
@limiter.limit("10/minute")
async def update_consumable(request: Request, consumable_id: str, req: ConsumableRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin"]))):
    result = await db.execute(update(Consumable).where(Consumable.id == consumable_id).values(name=req.name, unit=req.unit))
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(404, "Расходник не найден")
    return await get_consumable(request, consumable_id, db)

@router.delete("/{consumable_id}")
@limiter.limit("10/minute")
async def delete_consumable(request: Request, consumable_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin"]))):
    await db.execute(delete(ServiceConsumable).where(ServiceConsumable.consumableId == consumable_id))
    result = await db.execute(delete(Consumable).where(Consumable.id == consumable_id))
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(404, "Расходник не найден")
    return {"ok": True}

@router.post("/{consumable_id}/refill", response_model=ConsumableResponse)
@limiter.limit("10/minute")
async def refill_consumable(request: Request, consumable_id: str, req: RefillRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin", "washer"]))):
    result = await db.execute(select(Consumable).where(Consumable.id == consumable_id))
    consumable = result.scalar_one_or_none()
    if not consumable:
        raise HTTPException(404, "Расходник не найден")
    old_stock = consumable.currentStock
    consumable.currentStock += req.amount
    db.add(ConsumableRefillLog(
        consumableId=consumable_id,
        amount=req.amount,
        oldStock=old_stock,
        newStock=consumable.currentStock,
        refilledBy=current_user.username,
        timestamp=datetime.now().isoformat(),
    ))
    await db.commit()
    await db.refresh(consumable)
    return consumable

@router.get("/{consumable_id}/refill-history")
@limiter.limit("60/minute")
async def get_refill_history(request: Request, consumable_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin", "washer"]))):
    result = await db.execute(
        select(ConsumableRefillLog)
        .where(ConsumableRefillLog.consumableId == consumable_id)
        .order_by(ConsumableRefillLog.timestamp.desc())
    )
    logs = result.scalars().all()
    return [
        {
            "id": log.id,
            "amount": log.amount,
            "oldStock": log.oldStock,
            "newStock": log.newStock,
            "refilledBy": log.refilledBy,
            "timestamp": log.timestamp,
        }
        for log in logs
    ]

@router.get("/{consumable_id}/forecast")
@limiter.limit("60/minute")
async def get_consumable_forecast(request: Request, consumable_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin", "washer"]))):
    res = await db.execute(select(Consumable).where(Consumable.id == consumable_id))
    consumable = res.scalar_one_or_none()
    if not consumable:
        raise HTTPException(404, "Расходник не найден")

    thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()
    usage_res = await db.execute(
        select(func.coalesce(func.sum(ConsumableUsageLog.quantityUsed), 0.0))
        .where(
            ConsumableUsageLog.consumableId == consumable_id,
            ConsumableUsageLog.timestamp >= thirty_days_ago,
        )
    )
    total_used_30d = usage_res.scalar() or 0.0

    avg_daily = total_used_30d / 30.0 if total_used_30d > 0 else 0.0
    days_left = None
    if avg_daily > 0:
        days_left = consumable.currentStock / avg_daily

    target = consumable.minStock * 3
    to_buy = max(0.0, target - consumable.currentStock)

    return {
        "currentStock": consumable.currentStock,
        "minStock": consumable.minStock,
        "targetStock": target,
        "avgDailyUsage": round(avg_daily, 2),
        "daysLeft": round(days_left, 1) if days_left is not None else None,
        "suggestedPurchase": round(to_buy, 1),
        "unit": consumable.unit,
    }

@router.post("/service-link", response_model=ServiceConsumableResponse)
@limiter.limit("10/minute")
async def link_consumable_to_service(request: Request, req: ServiceConsumableRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin"]))):
    res_s = await db.execute(select(Service).where(Service.id == req.serviceId))
    if not res_s.scalar_one_or_none():
        raise HTTPException(404, f"Услуга с id={req.serviceId} не найдена")

    res_c = await db.execute(select(Consumable).where(Consumable.id == req.consumableId))
    if not res_c.scalar_one_or_none():
        raise HTTPException(404, f"Расходник с id={req.consumableId} не найден")

    existing = await db.execute(select(ServiceConsumable).where(ServiceConsumable.serviceId == req.serviceId, ServiceConsumable.consumableId == req.consumableId))
    link = existing.scalar_one_or_none()
    if link:
        link.quantity_per_service = req.quantity_per_service
    else:
        db.add(ServiceConsumable(serviceId=req.serviceId, consumableId=req.consumableId, quantity_per_service=req.quantity_per_service))
    
    await db.commit()
    return {"serviceId": req.serviceId, "consumableId": req.consumableId, "quantity_per_service": req.quantity_per_service}

@router.delete("/service-link/{service_id}/{consumable_id}")
@limiter.limit("10/minute")
async def unlink_consumable_from_service(request: Request, service_id: str, consumable_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(check_roles(["admin"]))):
    result = await db.execute(delete(ServiceConsumable).where(ServiceConsumable.serviceId == service_id, ServiceConsumable.consumableId == consumable_id))
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(404, "Связь расходника и услуги не найдена")
    return {"ok": True}
