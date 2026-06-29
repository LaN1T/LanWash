import asyncio
import io
import json
import uuid
from datetime import datetime, time, timedelta

from sqlalchemy.ext.asyncio import AsyncSession

from core.redis_client import get_redis
from models import (
    Consumable,
    ConsumableRefillLog,
    ServiceConsumable,
    WashTypeConsumable,
)
from repositories.consumable import ConsumableRepository
from repositories.consumable_refill_log import ConsumableRefillLogRepository
from repositories.consumable_usage_log import ConsumableUsageLogRepository
from repositories.service import ServiceRepository
from repositories.service_consumable import ServiceConsumableRepository
from repositories.wash_type import WashTypeRepository
from repositories.wash_type_consumable import WashTypeConsumableRepository
from schemas import (
    ConsumableRequest,
    RefillRequest,
    ServiceConsumableRequest,
    WashTypeConsumableRequest,
)
from services.inventory_forecast_service import generate_inventory_forecast

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ConsumableNotFoundError(Exception):
    pass


class ConsumablesService:
    """Business logic for consumables management."""

    def __init__(self, db: AsyncSession) -> None:
        self._db = db
        self._consumables = ConsumableRepository(db)
        self._services = ServiceRepository(db)
        self._service_consumables = ServiceConsumableRepository(db)
        self._wash_types = WashTypeRepository(db)
        self._wash_type_consumables = WashTypeConsumableRepository(db)
        self._refill_logs = ConsumableRefillLogRepository(db)
        self._usage_logs = ConsumableUsageLogRepository(db)

    async def get_all_consumables(self) -> list[Consumable]:
        return await self._consumables.list_all(order_by=Consumable.name.asc())

    async def get_consumables_by_service(
        self, service_id: str
    ) -> list[ServiceConsumable]:
        return await self._service_consumables.list_by_service(service_id)

    async def get_all_service_consumable_links(self) -> list[ServiceConsumable]:
        return await self._service_consumables.list_all()

    async def create_consumable(self, req: ConsumableRequest) -> Consumable:
        new_consumable = Consumable(id=str(uuid.uuid4()), name=req.name, unit=req.unit)
        await self._consumables.add(new_consumable)
        await self._db.commit()
        await self._db.refresh(new_consumable)
        return new_consumable

    async def get_low_stock_alerts(self) -> list[Consumable]:
        return await self._consumables.list_low_stock_alerts()

    _FORECAST_CACHE_TTL_SECONDS = 300

    async def get_inventory_forecast(self):
        cache_key = "inventory:forecast"
        try:
            redis = get_redis()
            if redis:
                cached = await redis.get(cache_key)
                if cached:
                    return json.loads(cached)
        except Exception:
            pass

        forecast = await generate_inventory_forecast(self._db)
        data = forecast.model_dump()

        try:
            redis = get_redis()
            if redis:
                await redis.setex(
                    cache_key, self._FORECAST_CACHE_TTL_SECONDS, json.dumps(data)
                )
        except Exception:
            pass

        return data

    async def get_consumable(self, consumable_id: str) -> Consumable | None:
        return await self._consumables.get_by_id(consumable_id)

    async def update_consumable(
        self, consumable_id: str, req: ConsumableRequest
    ) -> Consumable | None:
        updated = await self._consumables.update_by_id(
            consumable_id, name=req.name, unit=req.unit
        )
        if updated is None:
            return None
        await self._db.commit()
        return updated

    async def delete_consumable(self, consumable_id: str) -> bool:
        await self._service_consumables.delete_by_consumable_id(consumable_id)
        deleted = await self._consumables.delete_by_id(consumable_id)
        await self._db.commit()
        return deleted

    async def refill_consumable(
        self, consumable_id: str, req: RefillRequest, refilled_by: str
    ) -> Consumable | None:
        consumable = await self._consumables.get_by_id(consumable_id)
        if not consumable:
            return None
        old_stock = consumable.currentStock
        consumable.currentStock += req.amount
        await self._refill_logs.add(
            ConsumableRefillLog(
                consumableId=consumable_id,
                amount=req.amount,
                oldStock=old_stock,
                newStock=consumable.currentStock,
                refilledBy=refilled_by,
                timestamp=datetime.now(),
            )
        )
        await self._db.commit()
        await self._db.refresh(consumable)
        return consumable

    async def get_refill_history(self, consumable_id: str) -> list[dict]:
        logs = await self._refill_logs.list_by_consumable(consumable_id)
        return [
            {
                "id": log.id,
                "amount": log.amount,
                "oldStock": log.oldStock,
                "newStock": log.newStock,
                "refilledBy": log.refilledBy,
                "timestamp": log.timestamp,
            }
            for log in logs
        ]

    async def get_usage_history(self, consumable_id: str) -> list[dict]:
        rows = await self._usage_logs.list_by_consumable(consumable_id)
        return [
            {
                "consumableId": log.consumableId,
                "appointmentId": log.appointmentId,
                "quantityUsed": log.quantityUsed,
                "timestamp": log.timestamp,
                "appointmentDateTime": appt.dateTime,
                "carModel": appt.carModel,
                "carNumber": appt.carNumber,
                "washTypeId": appt.washTypeId,
            }
            for log, appt in rows
        ]

    async def get_history(
        self, consumable_id: str, history_type: str | None = None
    ) -> list[dict]:
        """Return a merged list of consumption and refill events for a consumable.

        Items are sorted by timestamp descending. If ``history_type`` is provided,
        only events of that type are returned.
        """
        items: list[dict] = []

        if history_type is None or history_type == "consumption":
            usage_rows = await self._usage_logs.list_by_consumable(consumable_id)
            for log, _appt in usage_rows:
                items.append(
                    {
                        "type": "consumption",
                        "id": log.id,
                        "appointmentId": log.appointmentId,
                        "quantity": log.quantityUsed,
                        "timestamp": log.timestamp,
                    }
                )

        if history_type is None or history_type == "refill":
            refill_logs = await self._refill_logs.list_by_consumable(consumable_id)
            for log in refill_logs:
                items.append(
                    {
                        "type": "refill",
                        "id": log.id,
                        "quantity": log.amount,
                        "timestamp": log.timestamp,
                    }
                )

        items.sort(key=lambda x: x["timestamp"], reverse=True)
        return items

    async def get_consumable_forecast(self, consumable_id: str) -> dict | None:
        consumable = await self._consumables.get_by_id(consumable_id)
        if not consumable:
            return None

        thirty_days_ago = datetime.now() - timedelta(days=30)
        total_used_30d = await self._usage_logs.sum_usage_since(
            consumable_id, thirty_days_ago
        )

        avg_daily = total_used_30d / 30.0 if total_used_30d > 0 else 0.0
        days_left = None
        if avg_daily > 0:
            days_left = consumable.currentStock / avg_daily

        target = consumable.minStock * 3
        to_buy = max(0.0, target - consumable.currentStock)

        return {
            "currentStock": consumable.currentStock,
            "minStock": consumable.minStock,
            "targetStock": target,
            "avgDailyUsage": round(avg_daily, 2),
            "daysLeft": round(days_left, 1) if days_left is not None else None,
            "suggestedPurchase": round(to_buy, 1),
            "unit": consumable.unit,
        }

    async def link_consumable_to_service(self, req: ServiceConsumableRequest) -> dict:
        service = await self._services.get_by_id(req.serviceId)
        if not service:
            raise ConsumableNotFoundError(f"Услуга с id={req.serviceId} не найдена")

        consumable = await self._consumables.get_by_id(req.consumableId)
        if not consumable:
            raise ConsumableNotFoundError(
                f"Расходник с id={req.consumableId} не найден"
            )

        link = await self._service_consumables.get_by_service_and_consumable(
            req.serviceId, req.consumableId
        )
        if link:
            link.quantity_per_service = req.quantity_per_service
        else:
            await self._service_consumables.add(
                ServiceConsumable(
                    serviceId=req.serviceId,
                    consumableId=req.consumableId,
                    quantity_per_service=req.quantity_per_service,
                )
            )

        await self._db.commit()
        return {
            "serviceId": req.serviceId,
            "consumableId": req.consumableId,
            "quantity_per_service": req.quantity_per_service,
        }

    async def unlink_consumable_from_service(
        self, service_id: str, consumable_id: str
    ) -> bool:
        deleted = await self._service_consumables.delete_by_service_and_consumable(
            service_id, consumable_id
        )
        await self._db.commit()
        return deleted

    async def get_consumables_by_wash_type(
        self, wash_type_id: str
    ) -> list[WashTypeConsumable]:
        return await self._wash_type_consumables.list_by_wash_type(wash_type_id)

    async def get_all_wash_type_consumable_links(self) -> list[WashTypeConsumable]:
        return await self._wash_type_consumables.list_all()

    async def link_consumable_to_wash_type(
        self, req: WashTypeConsumableRequest
    ) -> dict:
        wash_type = await self._wash_types.get_by_id(req.washTypeId)
        if not wash_type:
            raise ConsumableNotFoundError(f"Тип мойки с id={req.washTypeId} не найден")

        consumable = await self._consumables.get_by_id(req.consumableId)
        if not consumable:
            raise ConsumableNotFoundError(
                f"Расходник с id={req.consumableId} не найден"
            )

        link = await self._wash_type_consumables.get_by_wash_type_and_consumable(
            req.washTypeId, req.consumableId
        )
        if link:
            link.quantity_per_service = req.quantity_per_service
        else:
            await self._wash_type_consumables.add(
                WashTypeConsumable(
                    washTypeId=req.washTypeId,
                    consumableId=req.consumableId,
                    quantity_per_service=req.quantity_per_service,
                )
            )

        await self._db.commit()
        return {
            "washTypeId": req.washTypeId,
            "consumableId": req.consumableId,
            "quantity_per_service": req.quantity_per_service,
        }

    async def unlink_consumable_from_wash_type(
        self, wash_type_id: str, consumable_id: str
    ) -> bool:
        deleted = await self._wash_type_consumables.delete_by_wash_type_and_consumable(
            wash_type_id, consumable_id
        )
        await self._db.commit()
        return deleted

    # ─── Excel export / import helpers ───────────────────────────────────────

    def _create_workbook(self):
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl не установлен")
        return openpyxl.Workbook()

    @staticmethod
    def _style_header(cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1E88E5", end_color="1E88E5", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _auto_width(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _sanitize_excel(val):
        if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@", "\t", "\r"):
            return "'" + val
        return val

    _MAX_EXPORT_DAYS = 90

    async def export_consumables(
        self, date_from: str | None, date_to: str | None
    ) -> bytes:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl не установлен")

        today = datetime.now().date()
        if not date_to:
            date_to = today.isoformat()
        if not date_from:
            date_from = (
                datetime.fromisoformat(date_to).date() - timedelta(days=30)
            ).isoformat()

        try:
            from_date = datetime.fromisoformat(date_from).date()
            to_date = datetime.fromisoformat(date_to).date()
        except ValueError:
            raise ValueError(
                "date_from and date_to must be valid ISO dates (YYYY-MM-DD)"
            )

        if from_date > to_date:
            raise ValueError("date_from must not be later than date_to")
        if (to_date - from_date).days > self._MAX_EXPORT_DAYS:
            raise ValueError(
                f"Export range must not exceed {self._MAX_EXPORT_DAYS} days"
            )

        dt_start = datetime.combine(from_date, time.min)
        dt_end = datetime.combine(to_date, time.max)

        wb = self._create_workbook()
        wb.remove(wb.active)

        # 1. Лист "Остатки"
        ws_stock = wb.create_sheet("Остатки")
        headers_stock = [
            "ID",
            "Название",
            "Ед. изм.",
            "Текущий запас",
            "Мин. запас",
            "Статус",
        ]
        ws_stock.append(headers_stock)
        for cell in ws_stock[1]:
            self._style_header(cell)

        all_consumables = await self._consumables.list_all(
            order_by=Consumable.name.asc()
        )
        for c in all_consumables:
            status = "Низкий" if c.currentStock < c.minStock else "В норме"
            ws_stock.append(
                [
                    self._sanitize_excel(c.id),
                    self._sanitize_excel(c.name),
                    self._sanitize_excel(c.unit),
                    c.currentStock,
                    c.minStock,
                    self._sanitize_excel(status),
                ]
            )
        self._auto_width(ws_stock)

        # 2. Лист "Пополнения"
        ws_refill = wb.create_sheet("Пополнения")
        headers_refill = ["Расходник", "Количество", "Было", "Стало", "Кем", "Дата"]
        ws_refill.append(headers_refill)
        for cell in ws_refill[1]:
            self._style_header(cell)

        refill_logs = await self._refill_logs.list_by_date_range(dt_start, dt_end)
        refill_cids = {log.consumableId for log in refill_logs}
        cons_names: dict[str, str] = {}
        if refill_cids:
            consumables = await self._consumables.get_by_ids(list(refill_cids))
            cons_names = {c.id: c.name for c in consumables}

        for log in refill_logs:
            name = cons_names.get(log.consumableId, log.consumableId)
            ws_refill.append(
                [
                    self._sanitize_excel(name),
                    log.amount,
                    log.oldStock,
                    log.newStock,
                    self._sanitize_excel(log.refilledBy),
                    self._sanitize_excel(log.timestamp),
                ]
            )
        self._auto_width(ws_refill)

        # 3. Лист "Расход"
        ws_usage = wb.create_sheet("Расход")
        headers_usage = ["Расходник", "Ед. изм.", "Использовано", "Период"]
        ws_usage.append(headers_usage)
        for cell in ws_usage[1]:
            self._style_header(cell)

        usage_logs = await self._usage_logs.list_by_date_range(dt_start, dt_end)
        usage_cids = {log.consumableId for log in usage_logs}
        consumables_map: dict[str, tuple[str, str]] = {}
        if usage_cids:
            consumables = await self._consumables.get_by_ids(list(usage_cids))
            consumables_map = {c.id: (c.name, c.unit) for c in consumables}

        usage_sums: dict[str, float] = {}
        for log in usage_logs:
            usage_sums[log.consumableId] = (
                usage_sums.get(log.consumableId, 0.0) + log.quantityUsed
            )

        period_label = f"{date_from or '...'} - {date_to or '...'}"
        for cid, total in sorted(usage_sums.items(), key=lambda x: x[1], reverse=True):
            name, unit = consumables_map.get(cid, (cid, ""))
            ws_usage.append(
                [
                    self._sanitize_excel(name),
                    self._sanitize_excel(unit),
                    round(total, 2),
                    self._sanitize_excel(period_label),
                ]
            )
        self._auto_width(ws_usage)

        # 4. Лист "Прогноз"
        ws_forecast = wb.create_sheet("Прогноз")
        headers_forecast = [
            "Расходник",
            "Ед. изм.",
            "Текущий запас",
            "Средний расход/день",
            "Хватит на (дн)",
            "Реком. закупка",
        ]
        ws_forecast.append(headers_forecast)
        for cell in ws_forecast[1]:
            self._style_header(cell)

        all_consumables = await self._consumables.list_all()
        thirty_days_ago = datetime.now() - timedelta(days=30)
        usage_sums_map = await self._usage_logs.sum_usage_grouped_since(thirty_days_ago)

        for c in all_consumables:
            total_used = usage_sums_map.get(c.id, 0.0)
            avg_daily = total_used / 30.0 if total_used > 0 else 0.0
            days_left = round(c.currentStock / avg_daily, 1) if avg_daily > 0 else "—"
            target = c.minStock * 3
            to_buy = max(0.0, target - c.currentStock)
            ws_forecast.append(
                [
                    self._sanitize_excel(c.name),
                    self._sanitize_excel(c.unit),
                    c.currentStock,
                    round(avg_daily, 2),
                    self._sanitize_excel(str(days_left)),
                    round(to_buy, 1),
                ]
            )
        self._auto_width(ws_forecast)

        buf = io.BytesIO()
        await asyncio.to_thread(wb.save, buf)
        buf.seek(0)
        return buf.getvalue()

    async def generate_import_template(self) -> bytes:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl не установлен")

        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Пополнения"
        headers = ["name", "amount"]
        ws.append(headers)
        for cell in ws[1]:
            self._style_header(cell)
        ws.append(["Автошампунь", 10.0])
        ws.append(["Воск", 5.0])
        self._auto_width(ws)

        buf = io.BytesIO()
        await asyncio.to_thread(wb.save, buf)
        buf.seek(0)
        return buf.getvalue()

    async def import_refills(self, content: bytes, refilled_by: str) -> dict:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl не установлен")

        try:
            wb = await asyncio.to_thread(
                openpyxl.load_workbook,
                filename=io.BytesIO(content),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except Exception as e:
            raise ValueError(f"Не удалось открыть файл: {e}")

        ws = wb.active
        if not ws:
            raise ValueError("Файл пуст")

        header_row = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().strip() if h else "" for h in header_row]

        name_idx = None
        amount_idx = None
        for i, h in enumerate(header_lower):
            if h in ("name", "название", "расходник", "наименование"):
                name_idx = i
            if h in ("amount", "количество", "кол-во", "колво", "сумма"):
                amount_idx = i

        if name_idx is None or amount_idx is None:
            raise ValueError("Колонки 'name' и 'amount' не найдены в первой строке")

        succeeded = 0
        failed = 0
        errors: list[str] = []
        processed = 0

        all_consumables = await self._consumables.list_all()
        consumables_by_name = {c.name.lower(): c for c in all_consumables}

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = self._sanitize_excel(
                str(row[name_idx]).strip() if row[name_idx] else ""
            )
            amount_raw = row[amount_idx]
            if not name:
                continue
            processed += 1

            try:
                amount = float(amount_raw) if amount_raw is not None else 0.0
            except (ValueError, TypeError):
                failed += 1
                errors.append(
                    f"Строка {processed + 1}: '{amount_raw}' не является числом"
                )
                continue

            if amount <= 0:
                failed += 1
                errors.append(f"Строка {processed + 1}: количество должно быть > 0")
                continue

            consumable = consumables_by_name.get(name.lower())
            if not consumable:
                failed += 1
                errors.append(f"Строка {processed + 1}: расходник '{name}' не найден")
                continue

            old_stock = consumable.currentStock
            consumable.currentStock += amount
            await self._refill_logs.add(
                ConsumableRefillLog(
                    consumableId=consumable.id,
                    amount=amount,
                    oldStock=old_stock,
                    newStock=consumable.currentStock,
                    refilledBy=refilled_by,
                    timestamp=datetime.now(),
                )
            )
            succeeded += 1

        await self._db.commit()

        return {
            "processed": processed,
            "succeeded": succeeded,
            "failed": failed,
            "errors": errors[:20],
        }
